"""
Station data loading: reads Excel regression files and coordinate sheets.
"""

import json
import os

import openpyxl
from django.conf import settings

DATA_DIR = settings.DATA_DIR
CONFIG_PATH = os.path.join(DATA_DIR, "config.json")

# Station IDs to exclude (too far from target city to be useful)
EXCLUDED_STATION_IDS = {"50308", "50310", "50314", "50313", "55702"}

CITIES = {
    "Toronto":   {"label": "Toronto",   "lat": 43.7479, "lon": -79.2741},
    "Montreal":  {"label": "Montréal",  "lat": 45.5027, "lon": -73.6639},
    "Edmonton":  {"label": "Edmonton",  "lat": 53.5482, "lon": -113.3681},
    "Vancouver": {"label": "Vancouver", "lat": 49.3686, "lon": -123.2767},
}

DEMO_DATA = {
    "Toronto": {
        "60106": 85.0, "66201": 78.0, "65701": 72.0, "61201": 90.0,
        "60302": 65.0, "65401": 55.0, "60609": 30.0, "360291007": 20.0, "61502": 18.0,
    },
    "Montreal": {
        "54801": 80.0, "52001": 75.0, "50801": 68.0, "500070012": 55.0,
        "500070014": 50.0, "500070007": 45.0, "60106": 70.0, "60302": 40.0,
    },
    "Edmonton": {
        "92801": 90.0, "90302": 75.0, "94401": 65.0, "90304": 70.0,
        "91901": 55.0, "92901": 80.0,
    },
    "Vancouver": {
        "100316": 60.0, "100313": 55.0, "102301": 85.0, "102302": 80.0,
        "100304": 50.0, "100308": 45.0,
    },
}

# Cache loaded stations so we don't re-read Excel on every request
_station_cache = {}


def _find_col(headers, *candidates):
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        for c in candidates:
            if c.lower() in hl:
                return i
    return None


def load_stations(city_key):
    if city_key in _station_cache:
        return _station_cache[city_key]

    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return []

    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb["Included Stations"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return []

    headers = [str(h).strip() if h else "" for h in rows[1]]

    col_id    = _find_col(headers, "station id")
    col_city  = _find_col(headers, "city")
    col_dist  = _find_col(headers, "distance")
    col_dir   = _find_col(headers, "direction")
    col_tier  = _find_col(headers, "tier")
    col_slope = _find_col(headers, "slope")
    col_int   = _find_col(headers, "intercept")
    col_dtype = _find_col(headers, "data type")

    col_r = None
    for i, h in enumerate(headers):
        if h.strip() == "R":
            col_r = i
            break

    stations = []
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        if not sid:
            continue
        # Skip excluded stations
        if sid in EXCLUDED_STATION_IDS:
            continue
        city_name = str(row[col_city] or "")
        try:
            stations.append({
                "id": sid,
                "city_name": city_name,
                "distance": float(row[col_dist]) if row[col_dist] else 0,
                "direction": str(row[col_dir] or ""),
                "tier": int(str(row[col_tier]).replace("Tier", "").strip()) if row[col_tier] else 1,
                "R": float(row[col_r]) if col_r is not None and row[col_r] else 0,
                "slope": float(row[col_slope]) if row[col_slope] else 0,
                "intercept": float(row[col_int]) if row[col_int] else 0,
                "data_type": str(row[col_dtype] or "") if col_dtype is not None else "",
            })
        except (ValueError, TypeError):
            continue

    # Load lat/lon from All Stations Data sheet
    coord_map = _load_coords(city_key)
    for st in stations:
        c = coord_map.get(st["id"])
        if c:
            st["lat"] = c[0]
            st["lon"] = c[1]
        else:
            st["lat"] = None
            st["lon"] = None

    stations.sort(key=lambda s: (s["tier"], -s["distance"]))
    _station_cache[city_key] = stations
    return stations


def load_all_stations():
    """Load stations from all cities, tagging each with its target city."""
    if "_all" in _station_cache:
        return _station_cache["_all"]
    all_stations = []
    for city_key in CITIES:
        for st in load_stations(city_key):
            st_copy = dict(st)
            st_copy["target_city"] = city_key
            all_stations.append(st_copy)
    all_stations.sort(key=lambda s: (s["target_city"], s["tier"], -s["distance"]))
    _station_cache["_all"] = all_stations
    return all_stations


def get_all_demo_data():
    """Merge demo data from all cities into one dict."""
    merged = {}
    for city_data in DEMO_DATA.values():
        merged.update(city_data)
    return merged


def _load_coords(city_key):
    """Load lat/lon from 'All Stations Data' sheet. Returns {station_id: (lat, lon)}."""
    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return {}
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb["All Stations Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[1]]
    col_id = _find_col(headers, "station id")
    col_lat = _find_col(headers, "lat")
    col_lon = _find_col(headers, "lon")

    coords = {}
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        try:
            lat = float(row[col_lat])
            lon = float(row[col_lon])
            coords[sid] = (lat, lon)
        except (ValueError, TypeError):
            continue
    return coords
