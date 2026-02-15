"""
Core logic for the PM2.5 Early Warning System.
Loads Excel data, runs regression predictions, fetches live PM2.5 from PurpleAir.
"""

import json
import math
import os

import openpyxl
import requests
from django.conf import settings

DATA_DIR = settings.DATA_DIR
CONFIG_PATH = os.path.join(DATA_DIR, "config.json")

# ---------------------------------------------------------------------------
# Alert levels & colors (Toronto PM2.5 Methodology v3.0)
# ---------------------------------------------------------------------------
ALERT_LEVELS = [
    {"name": "LOW",       "min": 0,   "max": 20,    "hex": "#22c55e", "text_color": "black",
     "health": "No significant risk. No action required."},
    {"name": "MODERATE",  "min": 20,  "max": 60,    "hex": "#eab308", "text_color": "black",
     "health": "Sensitive groups (children, elderly, respiratory conditions) should reduce outdoor activity."},
    {"name": "HIGH",      "min": 60,  "max": 80,    "hex": "#f97316", "text_color": "black",
     "health": "General population affected. Reduce prolonged outdoor exertion. Use N95/KN95 mask outdoors."},
    {"name": "VERY HIGH", "min": 80,  "max": 120,   "hex": "#ef4444", "text_color": "white",
     "health": "Significant risk for all. Avoid outdoor exertion. Keep doors and windows closed."},
    {"name": "EXTREME",   "min": 120, "max": 1e9,   "hex": "#7f1d1d", "text_color": "white",
     "health": "Emergency conditions. Stay indoors. Close windows. Run HEPA filter. No indoor pollution sources."},
]

# ---------------------------------------------------------------------------
# Three-Rule Detection System (adapted from Toronto PM2.5 Methodology v3.0)
# ---------------------------------------------------------------------------
# Rule 1: Regional Station Alert - Any station > 40 µg/m³ triggers alert
# Rule 2: Distant Sequential Detection - Distant trigger + intermediate confirmation
# Rule 3: Corridor Detection - Upwind corridor stations for specific smoke sources

# Rule 1: Regional station trigger threshold
RULE1_TRIGGER = 40  # µg/m³ - Regional station exceeds this → evaluation begins

# Rule 2: Distant sequential detection thresholds
RULE2_DISTANT_TRIGGER = 35   # µg/m³ - Distant station (1000+ km) trigger
RULE2_INTERMEDIATE = 20      # µg/m³ - Intermediate station confirmation threshold

# Rule 3: Corridor detection threshold
RULE3_CORRIDOR_TRIGGER = 40  # µg/m³ - Corridor station trigger

# Common thresholds
CITY_ELEVATED_THRESHOLD = 20  # µg/m³ - City PM2.5 level that confirms smoke arrival
EVALUATION_WINDOW_HOURS = 120  # 5 days - Time window to check for city elevation
EVENT_COOLDOWN_HOURS = 168     # 7 days - Minimum separation between events
CONFIRMATION_WINDOW_HOURS = 96  # 4 days - Time for intermediate confirmation (Rule 2)

# Station IDs to exclude (too far from target city to be useful)
EXCLUDED_STATION_IDS = {"50308", "50310", "50314", "50313", "55702"}

CITIES = {
    "Toronto":   {"label": "Toronto",   "lat": 43.7479, "lon": -79.2741},
    "Montreal":  {"label": "Montréal",  "lat": 45.5027, "lon": -73.6639},
    "Edmonton":  {"label": "Edmonton",  "lat": 53.5482, "lon": -113.3681},
    "Vancouver": {"label": "Vancouver", "lat": 49.3686, "lon": -123.2767},
}

DEMO_DATA = {
    "Toronto": {
        "60106": 85.0, "66201": 78.0, "65701": 72.0, "61201": 90.0,
        "60302": 65.0, "65401": 55.0, "60609": 30.0, "360291007": 20.0, "61502": 18.0,
    },
    "Montreal": {
        "54801": 80.0, "52001": 75.0, "50801": 68.0, "500070012": 55.0,
        "500070014": 50.0, "500070007": 45.0, "60106": 70.0, "60302": 40.0,
    },
    "Edmonton": {
        "92801": 90.0, "90302": 75.0, "94401": 65.0, "90304": 70.0,
        "91901": 55.0, "92901": 80.0,
    },
    "Vancouver": {
        "100316": 60.0, "100313": 55.0, "102301": 85.0, "102302": 80.0,
        "100304": 50.0, "100308": 45.0,
    },
}

# ---------------------------------------------------------------------------
# Excel loading
# ---------------------------------------------------------------------------

def _find_col(headers, *candidates):
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        for c in candidates:
            if c.lower() in hl:
                return i
    return None


# Cache loaded stations so we don't re-read Excel on every request
_station_cache = {}


def load_stations(city_key):
    if city_key in _station_cache:
        return _station_cache[city_key]

    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return []

    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb["Included Stations"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return []

    headers = [str(h).strip() if h else "" for h in rows[1]]

    col_id    = _find_col(headers, "station id")
    col_city  = _find_col(headers, "city")
    col_dist  = _find_col(headers, "distance")
    col_dir   = _find_col(headers, "direction")
    col_tier  = _find_col(headers, "tier")
    col_slope = _find_col(headers, "slope")
    col_int   = _find_col(headers, "intercept")
    col_dtype = _find_col(headers, "data type")

    col_r = None
    for i, h in enumerate(headers):
        if h.strip() == "R":
            col_r = i
            break

    stations = []
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        if not sid:
            continue
        # Skip excluded stations
        if sid in EXCLUDED_STATION_IDS:
            continue
        city_name = str(row[col_city] or "")
        try:
            stations.append({
                "id": sid,
                "city_name": city_name,
                "distance": float(row[col_dist]) if row[col_dist] else 0,
                "direction": str(row[col_dir] or ""),
                "tier": int(str(row[col_tier]).replace("Tier", "").strip()) if row[col_tier] else 1,
                "R": float(row[col_r]) if col_r is not None and row[col_r] else 0,
                "slope": float(row[col_slope]) if row[col_slope] else 0,
                "intercept": float(row[col_int]) if row[col_int] else 0,
                "data_type": str(row[col_dtype] or "") if col_dtype is not None else "",
            })
        except (ValueError, TypeError):
            continue

    # Load lat/lon from All Stations Data sheet
    coord_map = _load_coords(city_key)
    for st in stations:
        c = coord_map.get(st["id"])
        if c:
            st["lat"] = c[0]
            st["lon"] = c[1]
        else:
            st["lat"] = None
            st["lon"] = None

    stations.sort(key=lambda s: (s["tier"], -s["distance"]))
    _station_cache[city_key] = stations
    return stations


def load_all_stations():
    """Load stations from all cities, tagging each with its target city."""
    if "_all" in _station_cache:
        return _station_cache["_all"]
    all_stations = []
    for city_key in CITIES:
        for st in load_stations(city_key):
            st_copy = dict(st)
            st_copy["target_city"] = city_key
            all_stations.append(st_copy)
    all_stations.sort(key=lambda s: (s["target_city"], s["tier"], -s["distance"]))
    _station_cache["_all"] = all_stations
    return all_stations


def get_all_demo_data():
    """Merge demo data from all cities into one dict."""
    merged = {}
    for city_data in DEMO_DATA.values():
        merged.update(city_data)
    return merged


def _load_coords(city_key):
    """Load lat/lon from 'All Stations Data' sheet. Returns {station_id: (lat, lon)}."""
    fn = os.path.join(DATA_DIR, f"{city_key}_PM25_EWS_Regression.xlsx")
    if not os.path.exists(fn):
        return {}
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb["All Stations Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[1]]
    col_id = _find_col(headers, "station id")
    col_lat = _find_col(headers, "lat")
    col_lon = _find_col(headers, "lon")

    coords = {}
    for row in rows[2:]:
        if row[col_id] is None:
            continue
        sid = str(row[col_id]).strip()
        try:
            lat = float(row[col_lat])
            lon = float(row[col_lon])
            coords[sid] = (lat, lon)
        except (ValueError, TypeError):
            continue
    return coords


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------

def get_alert_level(pm25):
    for lvl in reversed(ALERT_LEVELS):
        if pm25 >= lvl["min"]:
            return lvl
    return ALERT_LEVELS[0]


def lead_time_str(tier, dist):
    """Calculate estimated lead time based on station distance and tier.

    Lead times based on Toronto PM2.5 Methodology v3.0:
    - Distant stations (1000+ km): 15-92 hours (avg 15.3h)
    - Regional stations (100-600 km): 0-48 hours (varies by wind)
    - Corridor stations (300-500 km): 0-24 hours (often simultaneous)
    """
    if dist > 1000:
        return "24-72 hrs"
    if dist > 600:
        return "18-48 hrs"
    if dist > 400:
        return "12-36 hrs"
    if dist > 250:
        return "8-24 hrs"
    if dist > 150:
        return "4-18 hrs"
    return "2-12 hrs"


def _weighted_prediction(city_rows):
    """Calculate R-value weighted average prediction for a city.

    Stations with higher R-values (better correlation) have more influence.
    Uses R² as weight to emphasize high-correlation stations even more.
    Falls back to simple average if no valid R-values.
    """
    weighted_sum = 0.0
    weight_total = 0.0

    for r in city_rows:
        R = r.get("R", 0)
        pred = r["predicted"]
        # Use R² as weight (squares emphasize high-R stations)
        # Minimum weight of 0.1 to include all stations somewhat
        weight = max(R * R, 0.1)
        weighted_sum += weight * pred
        weight_total += weight

    if weight_total > 0:
        return weighted_sum / weight_total
    # Fallback to simple average
    return sum(r["predicted"] for r in city_rows) / len(city_rows) if city_rows else 0


def evaluate(stations, readings, previous_readings=None):
    """Evaluate stations using 3-rule detection system.

    Three-Rule Detection (adapted from Toronto PM2.5 Methodology v3.0):
    - Rule 1: Regional Alert - Any station > 40 µg/m³ (immediate trigger)
    - Rule 2: Distant Sequential - Distant station > 35 µg/m³ + intermediate > 20 µg/m³
    - Rule 3: Corridor Detection - Upwind corridor station > 40 µg/m³

    previous_readings: dict of {station_id: pm25} from the previous hour,
                       used for Rule 2 (sequential confirmation).

    Predictions are weighted by R-value (correlation coefficient) so that
    more reliable stations have greater influence on city-level predictions.
    """
    if previous_readings is None:
        previous_readings = {}

    # Build per-station results
    results = []
    for st in stations:
        sid = st["id"]
        if sid not in readings:
            continue
        pm = readings[sid]
        pred = st["slope"] * pm + st["intercept"]
        lvl = get_alert_level(pred)
        results.append({
            "station": st["city_name"], "id": sid,
            "dist": st["distance"], "dir": st["direction"],
            "tier": st["tier"], "R": st["R"], "pm25": pm,
            "predicted": round(pred, 1),
            "level_name": lvl["name"], "level_hex": lvl["hex"],
            "level_text_color": lvl["text_color"], "health": lvl["health"],
            "lead": lead_time_str(st["tier"], st["distance"]),
            "target_city": st.get("target_city", ""),
        })
    results.sort(key=lambda x: x["predicted"], reverse=True)

    # --- City-level alert determination using 3-rule system ---
    # Group results by target city
    city_results = {}
    for r in results:
        city = r.get("target_city", "")
        city_results.setdefault(city, []).append(r)

    city_alerts = {}
    for city, city_rows in city_results.items():
        alert_triggered = False
        trigger_rule = None
        trigger_stations = []

        # Calculate R-weighted prediction for this city
        weighted_pred = _weighted_prediction(city_rows)
        max_predicted = max((r["predicted"] for r in city_rows), default=0)

        # Categorize stations by distance (Tier indicates distance range)
        # Tier 1: 100-600 km (regional) - Rule 1
        # Tier 2+: Varies by city - can be distant or corridor
        regional_stations = [r for r in city_rows if r["tier"] == 1 and r["dist"] <= 600]
        distant_stations = [r for r in city_rows if r["dist"] > 600]
        corridor_stations = [r for r in city_rows if r["tier"] >= 2 and r["dist"] <= 400]

        # ═══════════════════════════════════════════════════════════════════
        # Rule 1: Regional Station Alert (100-600 km)
        # Any regional station > 40 µg/m³ triggers immediate alert
        # ═══════════════════════════════════════════════════════════════════
        for r in regional_stations:
            if r["pm25"] >= RULE1_TRIGGER:
                alert_triggered = True
                trigger_rule = "rule1"
                trigger_stations.append(r["station"])
                break

        # ═══════════════════════════════════════════════════════════════════
        # Rule 2: Distant Sequential Detection (600+ km)
        # Distant station > 35 µg/m³ AND intermediate station > 20 µg/m³
        # Uses previous readings to confirm smoke is moving toward city
        # ═══════════════════════════════════════════════════════════════════
        if not alert_triggered and previous_readings:
            # Check for distant trigger stations (> 35 µg/m³)
            distant_triggers = [
                r for r in distant_stations
                if r["pm25"] >= RULE2_DISTANT_TRIGGER
            ]
            # Check for intermediate confirmation (> 20 µg/m³, sustained)
            intermediate_confirmed = [
                r for r in city_rows
                if 200 <= r["dist"] <= 600  # Intermediate range
                and r["pm25"] >= RULE2_INTERMEDIATE
                and previous_readings.get(r["id"], 0) >= RULE2_INTERMEDIATE
            ]
            # Need distant trigger AND intermediate confirmation
            if distant_triggers and intermediate_confirmed:
                alert_triggered = True
                trigger_rule = "rule2"
                trigger_stations = [distant_triggers[0]["station"], intermediate_confirmed[0]["station"]]

        # ═══════════════════════════════════════════════════════════════════
        # Rule 3: Corridor Detection (upwind stations < 400 km)
        # Corridor station > 40 µg/m³ indicates smoke entering from
        # specific direction (e.g., NE for Quebec smoke, NW for Alberta)
        # ═══════════════════════════════════════════════════════════════════
        if not alert_triggered:
            for r in corridor_stations:
                if r["pm25"] >= RULE3_CORRIDOR_TRIGGER:
                    alert_triggered = True
                    trigger_rule = "rule3"
                    trigger_stations.append(r["station"])
                    break

        # Use weighted prediction for city-level alert determination
        alert_prediction = weighted_pred

        if alert_triggered:
            lvl = get_alert_level(alert_prediction)
            # Only issue alert if predicted > 20 (above LOW threshold)
            if lvl["name"] != "LOW":
                city_alerts[city] = {
                    "alert": True,
                    "rule": trigger_rule,
                    "trigger_stations": trigger_stations,
                    "predicted_pm25": round(alert_prediction, 1),
                    "weighted_pm25": round(weighted_pred, 1),
                    "max_pm25": round(max_predicted, 1),
                    "level_name": lvl["name"],
                    "level_hex": lvl["hex"],
                    "level_text_color": lvl["text_color"],
                    "health": lvl["health"],
                }

        if city not in city_alerts:
            low_lvl = ALERT_LEVELS[0]
            city_alerts[city] = {
                "alert": False,
                "rule": None,
                "trigger_stations": [],
                "predicted_pm25": round(weighted_pred, 1),
                "weighted_pm25": round(weighted_pred, 1),
                "max_pm25": round(max_predicted, 1),
                "level_name": low_lvl["name"],
                "level_hex": low_lvl["hex"],
                "level_text_color": low_lvl["text_color"],
                "health": low_lvl["health"],
            }

    return {"stations": results, "city_alerts": city_alerts}


# ---------------------------------------------------------------------------
# WAQI (World Air Quality Index) — aqicn.org
# ---------------------------------------------------------------------------

WAQI_BASE = "https://api.waqi.info"

# US EPA PM2.5 AQI breakpoints for AQI → µg/m³ conversion
_AQI_BREAKPOINTS = [
    (0,   50,   0.0,   12.0),
    (51,  100,  12.1,  35.4),
    (101, 150,  35.5,  55.4),
    (151, 200,  55.5,  150.4),
    (201, 300,  150.5, 250.4),
    (301, 400,  250.5, 350.4),
    (401, 500,  350.5, 500.4),
]


def _aqi_to_ugm3(aqi):
    """Convert PM2.5 AQI value to µg/m³ using US EPA breakpoints."""
    if aqi <= 0:
        return 0.0
    for aqi_lo, aqi_hi, c_lo, c_hi in _AQI_BREAKPOINTS:
        if aqi_lo <= aqi <= aqi_hi:
            return round(((aqi - aqi_lo) * (c_hi - c_lo)) / (aqi_hi - aqi_lo) + c_lo, 1)
    # Above 500 AQI — linear extrapolation
    return round(aqi * 1.0, 1)


def load_config():
    # Prefer environment variable over config file
    if os.environ.get("WAQI_API_TOKEN"):
        return {"api_key": os.environ["WAQI_API_TOKEN"]}
    try:
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        pass
    return {}


def _haversine(lat1, lon1, lat2, lon2):
    """Distance in km between two lat/lon points."""
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _fetch_waqi_bbox(token, lat1, lng1, lat2, lng2):
    """Fetch WAQI stations within a bounding box. Returns list of station dicts."""
    try:
        resp = requests.get(
            f"{WAQI_BASE}/v2/map/bounds",
            params={
                "latlng": f"{lat1},{lng1},{lat2},{lng2}",
                "networks": "all",
                "token": token,
            },
            timeout=30,
        )
        if resp.status_code != 200:
            return []
        data = resp.json()
    except requests.RequestException:
        return []

    if data.get("status") != "ok":
        return []

    result = []
    for entry in data.get("data", []):
        try:
            aqi_val = entry.get("aqi")
            if aqi_val is None or aqi_val == "-" or int(aqi_val) < 0:
                continue
            lat = entry["lat"]
            lon = entry["lon"]
            pm25 = _aqi_to_ugm3(int(aqi_val))
            result.append({
                "lat": float(lat),
                "lon": float(lon),
                "pm25": pm25,
                "name": entry.get("station", {}).get("name", ""),
            })
        except (KeyError, TypeError, ValueError):
            continue
    return result


def fetch_latest_pm25(api_key, stations):
    """Fetch PM2.5 for stations using per-city WAQI bounding-box queries.

    Groups stations by target_city and makes one bounding-box request
    per city. WAQI returns AQI values which are converted to µg/m³.
    """
    # Only stations with coordinates
    with_coords = [s for s in stations if s.get("lat") and s.get("lon")]
    if not with_coords:
        return {}

    # Group stations by target city
    city_groups = {}
    for s in with_coords:
        city = s.get("target_city", "")
        city_groups.setdefault(city, []).append(s)

    readings = {}
    pad = 0.5  # ~55 km padding

    for city, city_stations in city_groups.items():
        lats = [s["lat"] for s in city_stations]
        lons = [s["lon"] for s in city_stations]
        # WAQI bbox: lat1,lng1 = SW corner, lat2,lng2 = NE corner
        lat1 = min(lats) - pad
        lng1 = min(lons) - pad
        lat2 = max(lats) + pad
        lng2 = max(lons) + pad

        waqi_stations = _fetch_waqi_bbox(api_key, lat1, lng1, lat2, lng2)
        if not waqi_stations:
            continue

        # Match each station to nearest WAQI station within 30 km
        for st in city_stations:
            best_dist = 30  # km max
            best_pm = None
            for ws in waqi_stations:
                d = _haversine(st["lat"], st["lon"], ws["lat"], ws["lon"])
                if d < best_dist:
                    best_dist = d
                    best_pm = ws["pm25"]
            if best_pm is not None:
                readings[st["id"]] = best_pm

    return readings
